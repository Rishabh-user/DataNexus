import asyncio
from pathlib import Path

from app.core.logging import get_logger
from app.extraction.base import BaseExtractor, ExtractionResult, TableData

logger = get_logger(__name__)


class ExcelExtractor(BaseExtractor):
    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls"]

    async def extract(self, file_path: Path) -> ExtractionResult:
        return await asyncio.to_thread(self._extract_sync, file_path)

    def _extract_sync(self, file_path: Path) -> ExtractionResult:
        from openpyxl import load_workbook

        result = ExtractionResult()
        all_text_parts = []

        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        result.metadata = {
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data = []
            for row in ws.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                if any(v.strip() for v in row_values):
                    rows_data.append(row_values)

            if not rows_data:
                continue

            headers = rows_data[0]
            data_rows = rows_data[1:] if len(rows_data) > 1 else []

            result.tables.append(
                TableData(
                    headers=headers,
                    rows=data_rows,
                    sheet_name=sheet_name,
                )
            )

            text_repr = f"Sheet: {sheet_name}\n"
            text_repr += " | ".join(headers) + "\n"
            for row in data_rows:
                text_repr += " | ".join(row) + "\n"
            all_text_parts.append(text_repr)

        wb.close()
        result.text = "\n\n".join(all_text_parts)
        result.metadata["word_count"] = len(result.text.split())
        return result
