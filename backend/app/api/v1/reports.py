import json
import re
import uuid
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, status
from fastapi.responses import FileResponse as FastAPIFileResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.dependencies import PaginationParams
from app.core.logging import get_logger
from app.core.security import get_current_user, require_roles
from app.models.report import Report
from app.models.user import User, UserRole
from app.schemas.report import ReportGenerateRequest, ReportResponse

logger = get_logger(__name__)


def _extract_json(text: str) -> dict:
    """Robustly extract JSON object from LLM response text.

    Handles: bare JSON, ```json fenced blocks, extra text around JSON.
    """
    text = text.strip()

    # Try 1: direct parse
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Try 2: extract from markdown code fence
    fence_match = re.search(r'```(?:json)?\s*\n?(.*?)```', text, re.DOTALL)
    if fence_match:
        try:
            return json.loads(fence_match.group(1).strip())
        except json.JSONDecodeError:
            pass

    # Try 3: find first { ... last } in text
    first_brace = text.find('{')
    last_brace = text.rfind('}')
    if first_brace != -1 and last_brace > first_brace:
        try:
            return json.loads(text[first_brace:last_brace + 1])
        except json.JSONDecodeError:
            pass

    raise ValueError(f"Could not extract valid JSON from LLM response: {text[:200]}")


class ExcelFromChatRequest(BaseModel):
    """Generate Excel from AI chat response content."""
    content: str  # The AI response text
    title: str = "DataNexus Report"
    source_file_ids: list[int] = []  # Optional: include raw data from these files

router = APIRouter(prefix="/reports", tags=["Reports"])


@router.get("/templates")
async def list_templates():
    """Return available PPT template metadata for the UI."""
    from app.ppt.styles import THEMES
    return [
        {
            "id": t.id,
            "name": t.name,
            "description": t.description,
            "preview_colors": t.preview_colors,
        }
        for t in THEMES.values()
    ]


# --- Auth helpers for download (supports both Bearer header and ?token= query param) ---
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer

_optional_bearer = HTTPBearer(auto_error=False)


async def _optional_current_user(
    credentials: HTTPAuthorizationCredentials | None = Depends(_optional_bearer),
    db: AsyncSession = Depends(get_db),
) -> User | None:
    """Like get_current_user but returns None instead of raising 403."""
    if credentials is None:
        return None
    try:
        from app.core.security import decode_token
        payload = decode_token(credentials.credentials)
        user_id = int(payload["sub"])
        result = await db.execute(select(User).where(User.id == user_id))
        return result.scalar_one_or_none()
    except Exception:
        return None


async def _user_from_token(token: str, db: AsyncSession) -> User | None:
    """Resolve a user from a raw JWT token string (query param auth)."""
    try:
        from app.core.security import decode_token
        payload = decode_token(token)
        user_id = int(payload["sub"])
        result = await db.execute(select(User).where(User.id == user_id))
        return result.scalar_one_or_none()
    except Exception:
        return None


async def _generate_report_inline(user_id: int, report_id: int, title: str, prompt: str, include_charts: bool, template_id: str = "corporate"):
    """Generate report in background without Celery."""
    from app.core.database import async_session_factory
    from app.ai.llm import generate_chat_response
    from app.ai.prompts import REPORT_PROMPT_TEMPLATE
    from app.ai.retriever import retrieve_context
    from app.core.config import settings
    from app.ppt.generator import generate_ppt as create_ppt

    try:
        async with async_session_factory() as db:
            # 1. Get the existing report record
            result = await db.execute(select(Report).where(Report.id == report_id))
            report = result.scalar_one_or_none()
            if not report:
                logger.error("Report not found: %d", report_id)
                return

            # 2. Mark as generating
            report.generation_status = "generating"
            await db.commit()

            # 3. Retrieve relevant context
            retrieval = await retrieve_context(db=db, query=prompt, user_id=user_id, top_k=10)
            data_context = retrieval["context"]
            if retrieval["structured_data"]:
                data_context += "\n\nStructured Data:\n" + retrieval["structured_data"]

            # 4. Generate slide structure via LLM
            llm_prompt = REPORT_PROMPT_TEMPLATE.format(topic=title, data=data_context[:15000])
            response = await generate_chat_response(
                system_prompt="You are a presentation slide generator. You MUST return ONLY a valid JSON object with a 'title' string and a 'slides' array. No markdown, no explanation, no code fences — just raw JSON.",
                user_prompt=llm_prompt,
            )

            # 5. Parse slide structure JSON (robust extraction)
            slide_data = _extract_json(response)

            # 6. Generate the PPT file
            output_path = settings.reports_path / f"report_{report.id}.pptx"
            create_ppt(slide_data, str(output_path), include_charts=include_charts, theme_id=template_id)

            # 7. Update the report record
            report.file_path = str(output_path)
            report.generation_status = "completed"
            await db.commit()

            slide_count = len(slide_data.get("slides", []))
            logger.info("Report generated: %d - %s (%d slides)", report_id, title, slide_count)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        logger.error("Report generation failed for %d: %s\n%s", report_id, str(e), tb)
        try:
            async with async_session_factory() as db:
                result = await db.execute(select(Report).where(Report.id == report_id))
                report = result.scalar_one_or_none()
                if report:
                    report.generation_status = "failed"
                    report.error_message = f"{str(e)[:300]}\n---\n{tb[-200:]}"[:500]
                    await db.commit()
        except Exception:
            pass


@router.post("/generate-ppt", response_model=ReportResponse, status_code=202)
async def generate_ppt(
    request: ReportGenerateRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.ANALYST, UserRole.ADMIN, UserRole.SUPERADMIN)),
):
    # Create a pending report record
    report = Report(
        user_id=current_user.id,
        title=request.title,
        prompt_used=request.prompt,
        generation_status="pending",
    )
    db.add(report)
    await db.flush()

    # Dispatch background task (no Celery needed)
    background_tasks.add_task(
        _generate_report_inline,
        current_user.id, report.id, request.title, request.prompt,
        request.include_charts, request.template_id
    )

    return ReportResponse.model_validate(report)


@router.get("", response_model=list[ReportResponse])
async def list_reports(
    pagination: PaginationParams = Depends(),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(Report)
        .where(Report.user_id == current_user.id)
        .order_by(Report.created_at.desc())
        .offset(pagination.skip)
        .limit(pagination.limit)
    )
    reports = result.scalars().all()
    return [ReportResponse.model_validate(r) for r in reports]


@router.get("/{report_id}", response_model=ReportResponse)
async def get_report(
    report_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(Report).where(
            Report.id == report_id, Report.user_id == current_user.id
        )
    )
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Report not found")
    return ReportResponse.model_validate(report)


@router.get("/{report_id}/download")
async def download_report(
    report_id: int,
    token: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User | None = Depends(_optional_current_user),
):
    # Support auth via query param for direct browser downloads
    user = current_user
    if user is None and token:
        user = await _user_from_token(token, db)
    if user is None:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authenticated")

    result = await db.execute(
        select(Report).where(
            Report.id == report_id, Report.user_id == user.id
        )
    )
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Report not found")

    if report.generation_status != "completed" or not report.file_path:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Report not ready for download",
        )

    file_path = Path(report.file_path)
    if not file_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report file not found on disk",
        )

    return FastAPIFileResponse(
        path=str(file_path),
        filename=f"{report.title}.pptx",
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
    )


# =============================================
# Excel generation from AI chat responses
# =============================================

def _parse_markdown_tables(text: str) -> list[list[list[str]]]:
    """Extract markdown tables from text. Returns list of tables, each table is list of rows."""
    tables = []
    lines = text.split('\n')
    current_table = []
    for line in lines:
        stripped = line.strip()
        if '|' in stripped and stripped.startswith('|'):
            # Skip separator rows (---|---|---)
            if re.match(r'^[\|\s\-:]+$', stripped):
                continue
            cells = [c.strip() for c in stripped.split('|')[1:-1]]  # Remove empty first/last
            if cells:
                current_table.append(cells)
        else:
            if current_table:
                tables.append(current_table)
                current_table = []
    if current_table:
        tables.append(current_table)
    return tables


def _parse_pointer_sections(text: str) -> list[dict]:
    """Parse bullet-point / pointer sections from AI response.

    Returns list of {heading, items: [{key, value}]}
    """
    sections = []
    current_section = None

    for line in text.split('\n'):
        stripped = line.strip()
        # Detect headings (## or ### or **HEADING**)
        heading_match = re.match(r'^#{1,4}\s+(.+)', stripped)
        if heading_match:
            if current_section:
                sections.append(current_section)
            current_section = {"heading": heading_match.group(1).strip('*').strip(), "items": []}
            continue

        # Detect bold-only lines as headings
        bold_heading = re.match(r'^\*\*([^*]+)\*\*\s*$', stripped)
        if bold_heading and not stripped.startswith('•') and not stripped.startswith('-'):
            if current_section:
                sections.append(current_section)
            current_section = {"heading": bold_heading.group(1).strip(), "items": []}
            continue

        # Detect bullet points: • **Key**: Value  or  - **Key**: Value  or  • Value
        bullet_match = re.match(r'^[•\-\*]\s+\*\*([^*]+)\*\*[:\s]+(.+)', stripped)
        if bullet_match:
            if not current_section:
                current_section = {"heading": "Data", "items": []}
            current_section["items"].append({
                "key": bullet_match.group(1).strip(),
                "value": bullet_match.group(2).strip()
            })
            continue

        # Plain bullet: • Some text
        plain_bullet = re.match(r'^[•\-\*]\s+(.+)', stripped)
        if plain_bullet:
            if not current_section:
                current_section = {"heading": "Data", "items": []}
            current_section["items"].append({
                "key": "",
                "value": plain_bullet.group(1).strip()
            })

    if current_section:
        sections.append(current_section)
    return sections


def _generate_excel_from_content(content: str, title: str, extra_data: list[dict] = None) -> Path:
    """Create a well-formatted Excel workbook from AI response content."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Colour palette ──────────────────────────────────────────────────────
    TEAL      = "0D9488"
    NAVY      = "002F6C"
    WHITE     = "FFFFFF"
    MUTED     = "6B7280"
    LIGHT_BG  = "F0FDF9"
    ALT_ROW   = "F8FAFC"
    BORD_CLR  = "D1D5DB"

    def _side(style="thin", color=BORD_CLR):
        return Side(style=style, color=color)

    def _border(style="thin", color=BORD_CLR):
        s = _side(style, color)
        return Border(left=s, right=s, top=s, bottom=s)

    def _fill(color):
        return PatternFill("solid", fgColor=color)

    def _font(bold=False, size=11, color="1E293B", italic=False):
        return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # ── Sheet 1 — Report ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Report"
    ws.sheet_properties.tabColor = TEAL
    ws.freeze_panes = "A4"

    # Title block
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = title
    c.font  = _font(bold=True, size=16, color=NAVY)
    c.fill  = _fill(LIGHT_BG)
    c.alignment = _align()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    c2 = ws["A2"]
    c2.value = f"Generated by DataNexus AI  ·  {datetime.now().strftime('%d %b %Y, %H:%M')}"
    c2.font  = _font(size=9, color=MUTED, italic=True)
    c2.alignment = _align()
    ws.row_dimensions[2].height = 16

    row = 4

    # ── Sections / bullets ──────────────────────────────────────────────────
    sections = _parse_pointer_sections(content)
    if sections:
        for section in sections:
            # Section heading
            ws.merge_cells(f"A{row}:F{row}")
            c = ws.cell(row=row, column=1, value=section["heading"])
            c.font = _font(bold=True, size=12, color=TEAL)
            c.fill = _fill(LIGHT_BG)
            c.alignment = _align()
            c.border = _border()
            ws.row_dimensions[row].height = 22
            row += 1

            items = section.get("items", [])
            has_keys = any(item["key"] for item in items)

            if has_keys:
                # Column header row
                for ci, lbl in enumerate(["Field", "Value"], 1):
                    hc = ws.cell(row=row, column=ci, value=lbl)
                    hc.font  = _font(bold=True, size=11, color=WHITE)
                    hc.fill  = _fill(TEAL)
                    hc.alignment = _align("center")
                    hc.border = _border("medium", TEAL)
                ws.row_dimensions[row].height = 20
                row += 1
                for ri, item in enumerate(items):
                    bg = ALT_ROW if ri % 2 else WHITE
                    kc = ws.cell(row=row, column=1, value=item["key"])
                    kc.font  = _font(bold=True)
                    kc.fill  = _fill(bg)
                    kc.border = _border()
                    kc.alignment = _align()
                    vc = ws.cell(row=row, column=2, value=item["value"])
                    vc.font  = _font()
                    vc.fill  = _fill(bg)
                    vc.border = _border()
                    vc.alignment = _align(wrap=True)
                    ws.row_dimensions[row].height = 18
                    row += 1
            else:
                for ri, item in enumerate(items):
                    bg = ALT_ROW if ri % 2 else WHITE
                    ws.merge_cells(f"A{row}:F{row}")
                    c = ws.cell(row=row, column=1, value=f"  •  {item['value']}")
                    c.font = _font()
                    c.fill = _fill(bg)
                    c.border = _border()
                    c.alignment = _align(wrap=True)
                    ws.row_dimensions[row].height = 18
                    row += 1
            row += 1  # blank gap

    # ── Markdown tables ─────────────────────────────────────────────────────
    tables = _parse_markdown_tables(content)
    for table in tables:
        if not table:
            continue
        row += 1
        for ti, t_row in enumerate(table):
            for ci, val in enumerate(t_row, 1):
                c = ws.cell(row=row, column=ci, value=val.strip("*"))
                if ti == 0:
                    c.font = _font(bold=True, size=11, color=WHITE)
                    c.fill = _fill(TEAL)
                    c.alignment = _align("center")
                    c.border = _border("medium", TEAL)
                else:
                    bg = ALT_ROW if ti % 2 == 0 else WHITE
                    c.font = _font()
                    c.fill = _fill(bg)
                    c.alignment = _align(wrap=True)
                    c.border = _border()
                    # Auto-detect numbers
                    try:
                        clean = val.replace(",", "").replace("%", "").strip()
                        num = float(clean)
                        c.value = num
                        if "%" in val:
                            c.number_format = "0.00%"
                        elif "." in val:
                            c.number_format = "#,##0.00"
                        else:
                            c.number_format = "#,##0"
                    except (ValueError, AttributeError):
                        pass
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1

    # ── Fallback raw text ────────────────────────────────────────────────────
    if not sections and not tables:
        for line in content.split("\n"):
            s = line.strip()
            if s:
                ws.merge_cells(f"A{row}:F{row}")
                c = ws.cell(row=row, column=1, value=s)
                c.font = _font()
                c.alignment = _align(wrap=True)
                ws.row_dimensions[row].height = 16
                row += 1

    # ── Auto-fit column widths ───────────────────────────────────────────────
    defaults = {1: 28, 2: 55, 3: 20, 4: 20, 5: 20, 6: 20}
    for col, dw in defaults.items():
        letter = get_column_letter(col)
        max_len = dw
        for r in range(1, row):
            v = ws.cell(row=r, column=col).value
            if v:
                max_len = max(max_len, min(len(str(v)), 70))
        ws.column_dimensions[letter].width = max_len + 2

    # ── Sheet 2 — Extracted Data (optional) ─────────────────────────────────
    if extra_data:
        ws2 = wb.create_sheet("Extracted Data")
        ws2.sheet_properties.tabColor = "0072BD"
        ws2.freeze_panes = "A2"

        hdrs = ["File", "Data Type", "Page", "Content"]
        col_ws2 = [25, 16, 8, 90]
        for ci, (h, w) in enumerate(zip(hdrs, col_ws2), 1):
            hc = ws2.cell(row=1, column=ci, value=h)
            hc.font = _font(bold=True, size=11, color=WHITE)
            hc.fill = _fill(NAVY)
            hc.alignment = _align("center")
            hc.border = _border("medium", NAVY)
            ws2.column_dimensions[get_column_letter(ci)].width = w
        ws2.row_dimensions[1].height = 22

        for ri, ed in enumerate(extra_data, 2):
            bg = ALT_ROW if ri % 2 else WHITE
            vals = [
                ed.get("filename", ""),
                ed.get("data_type", ""),
                str(ed.get("source_page", "") or ""),
                (ed.get("raw_text", "") or "")[:3000],
            ]
            for ci, val in enumerate(vals, 1):
                c = ws2.cell(row=ri, column=ci, value=val)
                c.font = _font()
                c.fill = _fill(bg)
                c.border = _border()
                c.alignment = _align(wrap=True) if ci == 4 else _align()
            ws2.row_dimensions[ri].height = 40

    # ── Save ─────────────────────────────────────────────────────────────────
    output_dir = Path("uploads") / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{uuid.uuid4().hex[:12]}_{title[:30].replace(' ', '_')}.xlsx"
    filepath = output_dir / filename
    wb.save(str(filepath))
    return filepath


@router.post("/generate-excel")
async def generate_excel_from_chat(
    request: ExcelFromChatRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.ANALYST, UserRole.ADMIN, UserRole.SUPERADMIN)),
):
    """Generate a downloadable Excel file from AI chat response content."""

    # Optionally fetch raw extracted data from source files
    extra_data = []
    if request.source_file_ids:
        from app.ai.vector_store import _get_visible_user_ids
        from app.models.extracted_data import ExtractedData
        from app.models.file import File

        visible_ids = await _get_visible_user_ids(db, current_user.id)

        for fid in request.source_file_ids[:10]:  # Max 10 files
            # Verify file is accessible (own or team member's file)
            result = await db.execute(
                select(File).where(File.id == fid, File.user_id.in_(visible_ids))
            )
            f = result.scalar_one_or_none()
            if not f:
                continue

            result = await db.execute(
                select(ExtractedData).where(ExtractedData.file_id == fid)
            )
            for ed in result.scalars().all():
                extra_data.append({
                    "filename": f.filename,
                    "data_type": ed.data_type,
                    "source_page": ed.source_page,
                    "raw_text": ed.raw_text,
                    "structured_data": ed.structured_data,
                })

    filepath = _generate_excel_from_content(request.content, request.title, extra_data)

    return FastAPIFileResponse(
        path=str(filepath),
        filename=f"{request.title}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
